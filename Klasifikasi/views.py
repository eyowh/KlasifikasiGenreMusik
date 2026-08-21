import os
import io
import json
import joblib
import pickle
import tempfile
import warnings
import numpy as np
from pathlib import Path
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from .models import CustomUser
from django.contrib import messages


from .decorators import role_required, login_and_active_required, get_user_role

# ─── Load model artifacts on startup ───────────────────────────────────────────
MODEL_DIR = Path(__file__).resolve().parent / 'model'


def load_model_artifacts():
    try:
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            model = joblib.load(MODEL_DIR / 'svm_model.pkl')
            scaler = joblib.load(MODEL_DIR / 'scaler.pkl')
            label_encoder = joblib.load(MODEL_DIR / 'label_encoder.pkl')
        with open(MODEL_DIR / 'nama_fitur.pkl', 'rb') as f:
            feature_names = pickle.load(f)
        return model, scaler, label_encoder, feature_names
    except Exception as e:
        print(f"[WARNING] Gagal memuat model: {e}")
        return None, None, None, None


svm_model, scaler, label_encoder, feature_names = load_model_artifacts()

# ─── Supported audio formats ────────────────────────────────────────────────────
SUPPORTED_FORMATS = {
    '.mp3', '.wav', '.flac', '.ogg', '.m4a', '.aac',
    '.wma', '.aiff', '.aif', '.opus', '.mp4', '.webm',
    '.ape', '.wv', '.tta', '.mpc', '.spx', '.amr'
}

# ─── Audio Feature Extraction ───────────────────────────────────────────────────
def _get_ffmpeg_exe():
    """Cari executable FFmpeg dari berbagai sumber."""
    import shutil
    ffmpeg = shutil.which('ffmpeg') or shutil.which('ffmpeg.exe')
    if ffmpeg:
        return ffmpeg
    try:
        import imageio_ffmpeg
        exe = imageio_ffmpeg.get_ffmpeg_exe()
        if exe and os.path.exists(exe):
            return exe
    except Exception:
        pass
    return None


def _load_audio_robust(file_path, sr=22050, duration=30.0):
    """
    Load audio dengan multiple fallback strategy dan offset dinamis (tengah lagu):
    1. soundfile langsung (WAV, FLAC, OGG, AIFF)
    2. librosa default backend
    3. FFmpeg subprocess → konversi ke WAV (semua format: MP3, M4A, AAC, WMA, dll)
    """
    import librosa, warnings
    warnings.filterwarnings('ignore')

    ext = Path(file_path).suffix.lower()

    total_duration = 30.0
    try:
        total_duration = librosa.get_duration(path=file_path)
    except Exception:
        pass

    if total_duration > 60.0:
        offset = (total_duration - duration) / 2.0
    elif total_duration > duration:
        offset = total_duration - duration
    else:
        offset = 0.0

    # --- Layer 1: soundfile langsung ---
    soundfile_formats = {'.wav', '.flac', '.ogg', '.aiff', '.aif', '.opus'}
    if ext in soundfile_formats:
        try:
            import soundfile as sf
            file_info = sf.info(file_path)
            file_sr = file_info.samplerate
            total_frames = file_info.frames
            start_frame = int(offset * file_sr)
            stop_frame = int((offset + duration) * file_sr)
            if stop_frame > total_frames:
                stop_frame = total_frames
            data, file_sr = sf.read(file_path, start=start_frame, stop=stop_frame, always_2d=True)
            y = data.mean(axis=1).astype(np.float32)
            if file_sr != sr:
                y = librosa.resample(y, orig_sr=file_sr, target_sr=sr)
            y = y[:int(sr * duration)]
            if len(y) > 0:
                return y, sr
        except Exception as e1:
            print(f"[soundfile layer failed]: {e1}")

    # --- Layer 2: librosa default ---
    try:
        y, _ = librosa.load(file_path, sr=sr, offset=offset, duration=duration, mono=True,
                            res_type='kaiser_fast')
        if len(y) > 0:
            return y, sr
    except Exception as e2:
        print(f"[librosa direct layer failed]: {e2}")

    # --- Layer 3: FFmpeg subprocess → WAV ---
    ffmpeg_exe = _get_ffmpeg_exe()
    if ffmpeg_exe:
        import subprocess, tempfile as _tmplib
        tmp_wav = None
        try:
            tmp_wav = _tmplib.mktemp(suffix='.wav')
            cmd = [
                ffmpeg_exe, '-y',
                '-ss', f"{offset:.3f}",
                '-i', str(file_path),
                '-ar', str(sr),
                '-ac', '1',
                '-t', str(duration),
                '-vn',
                '-f', 'wav',
                tmp_wav
            ]
            result = subprocess.run(cmd, capture_output=True, timeout=60)
            if result.returncode == 0 and os.path.exists(tmp_wav):
                y, _ = librosa.load(tmp_wav, sr=sr, mono=True, res_type='kaiser_fast')
                if len(y) > 0:
                    return y, sr
        except Exception as e3:
            print(f"[ffmpeg layer failed]: {e3}")
        finally:
            if tmp_wav and os.path.exists(tmp_wav):
                try:
                    os.unlink(tmp_wav)
                except Exception:
                    pass
    else:
        print("[ffmpeg layer skipped]: FFmpeg tidak ditemukan")

    raise ValueError(
        f"Tidak dapat membaca file audio format '{ext}'. "
        "Untuk MP3/M4A/WMA/AAC, FFmpeg sudah tersedia (imageio-ffmpeg). "
        "Jika masih error, coba format WAV atau FLAC yang tidak memerlukan FFmpeg."
    )


def extract_features(file_path):
    """Ekstraksi 43 fitur audio menggunakan librosa — sesuai dengan model training."""
    try:
        import librosa
        warnings.filterwarnings('ignore')

        y, sr = _load_audio_robust(file_path, sr=22050, duration=30.0)

        if len(y) == 0:
            raise ValueError("File audio kosong atau tidak bisa dibaca.")

        features = {}

        # MFCC (13 × mean + std = 26 fitur)
        mfcc = librosa.feature.mfcc(y=y, sr=sr, n_mfcc=13)
        for i in range(13):
            features[f'mfcc_{i+1}_mean'] = float(np.mean(mfcc[i]))
            features[f'mfcc_{i+1}_std'] = float(np.std(mfcc[i]))

        # Chroma STFT
        chroma = librosa.feature.chroma_stft(y=y, sr=sr)
        features['chroma_mean'] = float(np.mean(chroma))
        features['chroma_std'] = float(np.std(chroma))

        # Spectral Centroid
        centroid = librosa.feature.spectral_centroid(y=y, sr=sr)
        features['spectral_centroid_mean'] = float(np.mean(centroid))
        features['spectral_centroid_std'] = float(np.std(centroid))

        # Spectral Bandwidth
        bandwidth = librosa.feature.spectral_bandwidth(y=y, sr=sr)
        features['spectral_bandwidth_mean'] = float(np.mean(bandwidth))
        features['spectral_bandwidth_std'] = float(np.std(bandwidth))

        # Spectral Rolloff
        rolloff = librosa.feature.spectral_rolloff(y=y, sr=sr)
        features['spectral_rolloff_mean'] = float(np.mean(rolloff))
        features['spectral_rolloff_std'] = float(np.std(rolloff))

        # Zero Crossing Rate
        zcr = librosa.feature.zero_crossing_rate(y)
        features['zcr_mean'] = float(np.mean(zcr))
        features['zcr_std'] = float(np.std(zcr))

        # RMS Energy
        rms = librosa.feature.rms(y=y)
        features['rms_mean'] = float(np.mean(rms))
        features['rms_std'] = float(np.std(rms))

        # Spectral Contrast
        contrast = librosa.feature.spectral_contrast(y=y, sr=sr)
        features['spectral_contrast_mean'] = float(np.mean(contrast))
        features['spectral_contrast_std'] = float(np.std(contrast))

        # Tonnetz
        try:
            tonnetz = librosa.feature.tonnetz(y=librosa.effects.harmonic(y), sr=sr)
            features['tonnetz_mean'] = float(np.mean(tonnetz))
            features['tonnetz_std'] = float(np.std(tonnetz))
        except Exception:
            features['tonnetz_mean'] = 0.0
            features['tonnetz_std'] = 0.0

        # Tempo
        try:
            tempo, _ = librosa.beat.beat_track(y=y, sr=sr)
            features['tempo'] = float(tempo) if np.isscalar(tempo) else float(tempo[0])
        except Exception:
            features['tempo'] = 0.0

        return features

    except ImportError:
        raise ImportError("librosa tidak terinstall. Jalankan: pip install librosa soundfile")
    except Exception as e:
        raise Exception(str(e))


def build_feature_vector(features_dict):
    """Susun fitur sesuai urutan yang digunakan saat training."""
    if feature_names is not None:
        try:
            feature_names_list = list(feature_names)
            vector = [features_dict.get(name, 0.0) for name in feature_names_list]
            return np.array(vector).reshape(1, -1), feature_names_list
        except Exception:
            pass
    names = list(features_dict.keys())
    vector = [features_dict[n] for n in names]
    return np.array(vector).reshape(1, -1), names


# ─── Genre Info ─────────────────────────────────────────────────────────────────
GENRE_INFO = {
    'blues':     {'color': '#4169E1', 'icon': '🎸', 'desc': 'Genre musik yang berasal dari komunitas Afrika-Amerika, ditandai dengan penggunaan blues scale dan call-and-response pattern.'},
    'classical': {'color': '#9B59B6', 'icon': '🎻', 'desc': 'Musik orkestra dari era Baroque hingga Romantic, dikenal dengan komposisi terstruktur dan instrumentasi orkestra.'},
    'country':   {'color': '#F39C12', 'icon': '🤠', 'desc': 'Genre musik Amerika yang berakar dari folk, ditandai dengan gitar akustik, fiddle, dan lirik naratif.'},
    'disco':     {'color': '#E91E63', 'icon': '🕺', 'desc': 'Genre dance music dari era 1970an, dikenal dengan beat 4/4 yang kuat, bass lines, dan string arrangements.'},
    'hiphop':    {'color': '#FF5722', 'icon': '🎤', 'desc': 'Genre yang lahir dari budaya hip-hop, ditandai dengan rapping, sampling, dan heavy beats.'},
    'jazz':      {'color': '#00BCD4', 'icon': '🎷', 'desc': 'Genre improvisasi yang kompleks, kaya dengan harmoni jazz, syncopated rhythms, dan eksplorasi melodik.'},
    'metal':     {'color': '#607D8B', 'icon': '🤘', 'desc': 'Genre rock yang distorsif dan keras, ditandai dengan heavy guitar riffs, drumming cepat, dan vokalisasi powerful.'},
    'pop':       {'color': '#FF4081', 'icon': '🎵', 'desc': 'Genre musik populer mainstream dengan melodi catchy, struktur verse-chorus, dan produksi yang polished.'},
    'reggae':    {'color': '#4CAF50', 'icon': '🌴', 'desc': 'Genre dari Jamaika, ditandai dengan offbeat chords, syncopated rhythms, dan bass lines yang dominan.'},
    'rock':      {'color': '#F44336', 'icon': '🎸', 'desc': 'Genre yang berpusat pada guitar, bass, dan drums, dengan pengaruh blues dan energi yang kuat.'},
}



# ═══════════════════════════════════════════════════════════════════════
# LANDING PAGE (public)
# ═══════════════════════════════════════════════════════════════════════

def landing_view(request):
    """Halaman landing publik — redirect ke dashboard jika sudah login."""
    if request.user.is_authenticated:
        return redirect('Klasifikasi:home')

    # Ambil data publik jika ada (statistik sederhana untuk ditampilkan di landing)
    try:
        from .models import RiwayatKlasifikasi
        total_klasifikasi = RiwayatKlasifikasi.objects.count()
    except Exception:
        total_klasifikasi = 0

    try:
        with open(MODEL_DIR / 'metrics.json', 'r') as f:
            metrics = json.load(f)
        accuracy = f"{metrics['overall']['accuracy'] * 100:.1f}"
    except Exception:
        accuracy = None

    return render(request, 'Klasifikasi/landing.html', {
        'total_klasifikasi': total_klasifikasi,
        'accuracy': accuracy,
    })


# ═══════════════════════════════════════════════════════════════════════
# AUTH VIEWS
# ═══════════════════════════════════════════════════════════════════════

def login_view(request):
    """Halaman login."""
    if request.user.is_authenticated:
        return redirect('Klasifikasi:home')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'Klasifikasi:home')
            return redirect(next_url)
        else:
            return render(request, 'Klasifikasi/login.html', {
                'error': 'Username atau password salah.',
                'username': username,
            })

    return render(request, 'Klasifikasi/login.html', {})


def logout_view(request):
    """Logout dan redirect ke landing."""
    logout(request)
    return redirect('Klasifikasi:landing')


def register_view(request):
    """Halaman registrasi publik — membuat akun pengguna_studio."""
    from .models import UserProfile, CustomUser
    if request.user.is_authenticated:
        return redirect('Klasifikasi:home')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        first_name = request.POST.get('first_name', '').strip()

        errors = {}
        if not username:
            errors['username'] = 'Username wajib diisi.'
        elif CustomUser.objects.filter(username=username).exists():
            errors['username'] = 'Username sudah digunakan.'
        if not password1:
            errors['password1'] = 'Password wajib diisi.'
        elif len(password1) < 8:
            errors['password1'] = 'Password minimal 8 karakter.'
        elif password1 != password2:
            errors['password2'] = 'Password tidak cocok.'

        if not errors:
            user = CustomUser.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
            )
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = 'pengguna_studio'
            profile.is_active = True
            profile.save()
            messages.success(request, f'Akun "{username}" berhasil dibuat. Silakan login.')
            return redirect('Klasifikasi:login')

        return render(request, 'Klasifikasi/register.html', {
            'errors': errors,
            'form_data': {'username': username, 'email': email, 'first_name': first_name},
        })

    return render(request, 'Klasifikasi/register.html', {})


def forbidden_view(request):
    """Halaman 403 — akses ditolak."""
    return render(request, 'Klasifikasi/403.html', {}, status=403)


# ═══════════════════════════════════════════════════════════════════════
# HOME / DASHBOARD
# ═══════════════════════════════════════════════════════════════════════

@login_and_active_required
def home_view(request):
    """Halaman utama / dashboard — konten berbeda per role."""
    from .models import RiwayatKlasifikasi

    role = get_user_role(request.user)

    with open(MODEL_DIR / 'metrics.json', 'r') as f:
        metrics = json.load(f)

    # Stats personal selalu dihitung
    riwayat_user = RiwayatKlasifikasi.objects.filter(user=request.user)
    total_klasifikasi_user = riwayat_user.count()
    genre_counts_user = {}
    for r in riwayat_user:
        genre_counts_user[r.genre_prediksi] = genre_counts_user.get(r.genre_prediksi, 0) + 1
    top_genre_user = max(genre_counts_user, key=genre_counts_user.get) if genre_counts_user else '-'

    # 5 klasifikasi terbaru untuk tabel dashboard
    riwayat_saya = []
    for r in riwayat_user.order_by('-waktu_klasifikasi')[:5]:
        genre_detail = GENRE_INFO.get(r.genre_prediksi.lower(), {})
        riwayat_saya.append({
            'nama_file': r.nama_file,
            'genre_prediksi': r.genre_prediksi,
            'genre_icon': r.genre_icon,
            'genre_color': genre_detail.get('color', '#888'),
            'confidence': r.confidence,
            'waktu': r.waktu_klasifikasi,
        })

    context = {
        'role': role,
        'accuracy': f"{metrics['overall']['accuracy'] * 100:.1f}",
        'jumlah_genre': len(metrics['per_class']),
        'jumlah_fitur': metrics['model_info']['jumlah_fitur'],
        'kernel': metrics['model_info']['kernel'],
        'model_info': metrics['model_info'],
        'total_klasifikasi_user': total_klasifikasi_user,
        'top_genre_user': top_genre_user,
        'genre_counts_user_json': json.dumps(genre_counts_user),
        'riwayat_saya': riwayat_saya,
    }

    # Extra stats untuk pengelola / admin
    if role in ('admin', 'pengelola_studio'):
        total_all = RiwayatKlasifikasi.objects.count()
        genre_counts_all = {}
        for r in RiwayatKlasifikasi.objects.all():
            genre_counts_all[r.genre_prediksi] = genre_counts_all.get(r.genre_prediksi, 0) + 1
        total_users = CustomUser.objects.count()

        context.update({
            'total_klasifikasi_all': total_all,
            'genre_counts_all_json': json.dumps(genre_counts_all),
            'total_users': total_users,
        })

    return render(request, 'Klasifikasi/home.html', context)



# ═══════════════════════════════════════════════════════════════════════
# KLASIFIKASI — upload & predict
# ═══════════════════════════════════════════════════════════════════════

@login_and_active_required
def klasifikasi_view(request):
    """Halaman klasifikasi genre musik — upload audio → prediksi."""
    from .models import RiwayatKlasifikasi, FiturAudio

    role = get_user_role(request.user)

    context = {
        'supported_formats': ', '.join(sorted(SUPPORTED_FORMATS)),
        'model_ready': svm_model is not None,
        'role': role,
    }

    if request.method == 'POST':
        audio_file = request.FILES.get('audio_file')

        if not audio_file:
            context['error'] = 'Tidak ada file yang diupload.'
            return render(request, 'Klasifikasi/klasifikasi.html', context)

        ext = Path(audio_file.name).suffix.lower()
        if ext not in SUPPORTED_FORMATS:
            context['error'] = f'Format "{ext}" tidak didukung.'
            return render(request, 'Klasifikasi/klasifikasi.html', context)

        if svm_model is None:
            context['error'] = 'Model SVM belum dimuat. Pastikan file model tersedia.'
            return render(request, 'Klasifikasi/klasifikasi.html', context)

        suffix = ext if ext else '.wav'
        tmp_file = None
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                for chunk in audio_file.chunks():
                    tmp.write(chunk)
                tmp_file = tmp.name

            features_dict = extract_features(tmp_file)
            feature_vector, used_feature_names = build_feature_vector(features_dict)
            feature_vector_scaled = scaler.transform(feature_vector)

            prediction = svm_model.predict(feature_vector_scaled)
            genre_raw = label_encoder.inverse_transform(prediction)[0]
            genre = genre_raw.lower()

            probabilities = []
            confidence = None
            try:
                proba = svm_model.predict_proba(feature_vector_scaled)[0]
                genre_labels = label_encoder.classes_
                proba_sorted = sorted(zip(genre_labels, proba), key=lambda x: x[1], reverse=True)
                probabilities = [
                    {
                        'genre': g,
                        'probability': round(float(p) * 100, 2),
                        'color': GENRE_INFO.get(g.lower(), {}).get('color', '#888'),
                        'icon': GENRE_INFO.get(g.lower(), {}).get('icon', '🎵'),
                    }
                    for g, p in proba_sorted
                ]
                confidence = round(float(max(proba)) * 100, 2)
            except AttributeError:
                confidence = None
                probabilities = []

            # Confidence level
            if confidence is not None:
                if confidence >= 60:
                    confidence_level = 'tinggi'
                elif confidence >= 40:
                    confidence_level = 'sedang'
                else:
                    confidence_level = 'rendah'
            else:
                confidence_level = None

            top_features = sorted(features_dict.items(), key=lambda x: abs(x[1]), reverse=True)[:10]
            genre_detail = GENRE_INFO.get(genre, {'color': '#888', 'icon': '🎵', 'desc': f'Genre musik {genre_raw}.'})
            filesize_str = f"{audio_file.size / 1024:.1f} KB" if audio_file.size < 1048576 else f"{audio_file.size / 1048576:.1f} MB"

            genre_db_desc = genre_detail.get('desc', '')

            # ── Simpan ke riwayat ──
            riwayat = RiwayatKlasifikasi.objects.create(
                user=request.user,
                nama_file=audio_file.name,
                ukuran_file=filesize_str,
                format_file=ext.upper().replace('.', ''),
                genre_prediksi=genre_raw,
                genre_icon=genre_detail.get('icon', '🎵'),
                confidence=confidence,
                top_features_json=json.dumps(top_features),
                probabilities_json=json.dumps(probabilities),
            )

            # ── Simpan fitur audio ke tb_fitur_audio ──
            try:
                FiturAudio.objects.create(
                    id_riwayat=riwayat,
                    mfcc_mean=features_dict.get('mfcc_1_mean', 0.0),
                    spectral_centroid=features_dict.get('spectral_centroid_mean', 0.0),
                    zero_crossing_rate=features_dict.get('zcr_mean', 0.0),
                    chroma_mean=features_dict.get('chroma_mean', 0.0),
                    tempo=features_dict.get('tempo', 0.0),
                )
            except Exception as fe:
                print(f"[WARNING] Gagal menyimpan FiturAudio: {fe}")

            context.update({
                'result': True,
                'genre': genre_raw,
                'genre_lower': genre,
                'genre_icon': genre_detail.get('icon', '🎵'),
                'genre_color': genre_detail.get('color', '#888'),
                'genre_desc': genre_db_desc,
                'confidence': confidence,
                'confidence_level': confidence_level,
                'probabilities': probabilities,
                'top3_probabilities': probabilities[:3],
                'top_features': top_features,
                'filename': audio_file.name,
                'filesize': filesize_str,
                # Feature values for display (digunakan oleh admin & pengelola)
                'feat_mfcc_mean': round(features_dict.get('mfcc_1_mean', 0.0), 4),
                'feat_spectral_centroid': round(features_dict.get('spectral_centroid_mean', 0.0), 2),
                'feat_zcr': round(features_dict.get('zcr_mean', 0.0), 6),
                'feat_chroma_mean': round(features_dict.get('chroma_mean', 0.0), 4),
                'feat_tempo': round(features_dict.get('tempo', 0.0), 2),
                'feat_spectral_bandwidth': round(features_dict.get('spectral_bandwidth_mean', 0.0), 2),
                'feat_rms': round(features_dict.get('rms_mean', 0.0), 6),
            })

        except ImportError as e:
            context['error'] = str(e)
        except Exception as e:
            context['error'] = f'Error saat memproses audio: {str(e)}'
        finally:
            if tmp_file and os.path.exists(tmp_file):
                try:
                    os.unlink(tmp_file)
                except Exception:
                    pass

    return render(request, 'Klasifikasi/klasifikasi.html', context)


# ═══════════════════════════════════════════════════════════════════════
# KINERJA MODEL
# ═══════════════════════════════════════════════════════════════════════

@role_required('admin', 'pengelola_studio')
def kinerja_view(request):
    """Halaman analisis kinerja model SVM."""
    with open(MODEL_DIR / 'metrics.json', 'r') as f:
        metrics = json.load(f)

    per_class_list = []
    for genre, vals in metrics['per_class'].items():
        f1 = vals['f1_score']

        # ── Interpretasi F1-score untuk pengguna awam ──
        if f1 >= 0.85:
            performance_label = 'Sangat Baik'
            performance_color = 'green'
            performance_desc  = 'Genre ini dikenali dengan sangat akurat oleh model.'
        elif f1 >= 0.70:
            performance_label = 'Baik'
            performance_color = 'yellow'
            performance_desc  = 'Genre ini cukup sering dikenali dengan benar, meski ada beberapa kesalahan.'
        else:
            performance_label = 'Perlu Ditingkatkan'
            performance_color = 'red'
            performance_desc  = 'Genre ini sering tertukar dengan genre lain — model masih kesulitan mengenalinya.'

        per_class_list.append({
            'genre': genre,
            'precision': vals['precision'],
            'recall': vals['recall'],
            'f1_score': f1,
            'support': vals['support'],
            'icon': GENRE_INFO.get(genre.lower(), {}).get('icon', '\U0001f3b5'),
            'color': GENRE_INFO.get(genre.lower(), {}).get('color', '#888'),
            'performance_label': performance_label,
            'performance_color': performance_color,
            'performance_desc':  performance_desc,
        })

    cv = metrics.get('cross_validation', {})

    # ── Interpretasi stabilitas std cross-validation ──
    cv_std_raw = cv.get('std', 0)
    if cv_std_raw <= 0.015:
        cv_std_level = 'sangat stabil'
    elif cv_std_raw <= 0.03:
        cv_std_level = 'cukup stabil'
    else:
        cv_std_level = 'kurang konsisten'

    cv_folds = [
        {'fold': i+1, 'score': cv.get(f'fold_{i+1}', 0), 'score_css': str(cv.get(f'fold_{i+1}', 0))}
        for i in range(5)
    ]

    # ── Temukan pasangan genre yang paling sering tertukar (di luar diagonal) ──
    worst_confusion_pair = None
    try:
        cm_data = metrics.get('confusion_matrix', {})
        cm_matrix = cm_data.get('matrix', [])
        cm_labels = cm_data.get('labels', [])
        max_off = -1
        for r_idx, row in enumerate(cm_matrix):
            for c_idx, val in enumerate(row):
                if r_idx != c_idx and val > max_off:
                    max_off = val
                    worst_confusion_pair = {
                        'actual': cm_labels[r_idx],
                        'predicted': cm_labels[c_idx],
                        'count': val,
                    }
    except Exception:
        worst_confusion_pair = None

    context = {
        'metrics': metrics,
        'per_class_list': per_class_list,
        'accuracy_pct': f"{metrics['overall']['accuracy'] * 100:.2f}",
        'precision_pct': f"{metrics['overall']['macro_precision'] * 100:.2f}",
        'recall_pct': f"{metrics['overall']['macro_recall'] * 100:.2f}",
        'f1_pct': f"{metrics['overall']['macro_f1'] * 100:.2f}",
        'cv_mean_pct': f"{cv.get('mean', 0) * 100:.2f}",
        'cv_std': f"{cv_std_raw:.4f}",
        'cv_std_level': cv_std_level,
        'cv_folds': cv_folds,
        'cv_folds_json': json.dumps([f['score'] for f in cv_folds]),
        'confusion_matrix_json': json.dumps(metrics['confusion_matrix']),
        'per_class_json': json.dumps(metrics['per_class']),
        'worst_confusion_pair': worst_confusion_pair,
        'laporan_gambar_cm': 'laporan/confusion_matrix.png',
        'laporan_gambar_cv': 'laporan/cross_validation.png',
    }
    return render(request, 'Klasifikasi/kinerja.html', context)



# ═══════════════════════════════════════════════════════════════════════
# RIWAYAT — personal (semua role)
# ═══════════════════════════════════════════════════════════════════════

@login_and_active_required
def riwayat_view(request):
    """
    Halaman riwayat klasifikasi.
    - pengguna_studio: hanya milik sendiri, tanpa tab.
    - admin / pengelola_studio: 2 tab — "Riwayat Saya" dan "Semua Riwayat".
      Tab aktif dikontrol via ?tab=semua (default: 'saya').
    """
    from .models import RiwayatKlasifikasi

    role = get_user_role(request.user)
    is_advanced = role in ('admin', 'pengelola_studio')

    # ── Tab aktif ──
    active_tab = request.GET.get('tab', 'saya')
    if not is_advanced:
        active_tab = 'saya'  # paksa tab saya untuk pengguna biasa

    # ══════════════════════════════════════════════════
    # TAB SAYA — riwayat milik sendiri
    # ══════════════════════════════════════════════════
    genre_filter_saya = request.GET.get('genre', '') if active_tab == 'saya' else ''
    qs_saya = RiwayatKlasifikasi.objects.filter(user=request.user)
    if genre_filter_saya:
        qs_saya = qs_saya.filter(genre_prediksi__iexact=genre_filter_saya)

    riwayat_saya = []
    for r in qs_saya:
        try:
            probabilities = json.loads(r.probabilities_json) if r.probabilities_json else []
        except Exception:
            probabilities = []
        try:
            top_features = json.loads(r.top_features_json) if r.top_features_json else []
        except Exception:
            top_features = []
        fitur = None
        try:
            fitur = r.fitur_audio
        except Exception:
            pass
        genre_detail = GENRE_INFO.get(r.genre_prediksi.lower(), {})
        riwayat_saya.append({
            'id': r.id_riwayat,
            'nama_file': r.nama_file,
            'ukuran_file': r.ukuran_file,
            'format_file': r.format_file,
            'genre_prediksi': r.genre_prediksi,
            'genre_icon': r.genre_icon,
            'genre_color': genre_detail.get('color', '#888'),
            'genre_desc': genre_detail.get('desc', ''),
            'confidence': r.confidence,
            'waktu': r.waktu_klasifikasi,
            'probabilities_json': json.dumps(probabilities),
            'top_features_json': json.dumps(top_features),
            'fitur': fitur,
        })

    total_saya = RiwayatKlasifikasi.objects.filter(user=request.user).count()
    genre_stats_saya = {}
    for r in RiwayatKlasifikasi.objects.filter(user=request.user):
        genre_stats_saya[r.genre_prediksi] = genre_stats_saya.get(r.genre_prediksi, 0) + 1
    raw_genres_saya = RiwayatKlasifikasi.objects.filter(user=request.user).values_list('genre_prediksi', flat=True)
    all_genres_saya = sorted(list(set(g.strip().lower() for g in raw_genres_saya if g)))

    # ══════════════════════════════════════════════════
    # TAB SEMUA — hanya untuk admin / pengelola
    # ══════════════════════════════════════════════════
    riwayat_semua = []
    total_semua = 0
    genre_stats_semua = {}
    all_genres_semua = []
    genre_filter_semua = ''
    user_filter_semua = ''
    date_from_semua = ''
    date_to_semua = ''

    if is_advanced:
        genre_filter_semua = request.GET.get('genre', '') if active_tab == 'semua' else ''
        user_filter_semua  = request.GET.get('user', '')  if active_tab == 'semua' else ''
        date_from_semua    = request.GET.get('date_from', '') if active_tab == 'semua' else ''
        date_to_semua      = request.GET.get('date_to', '')   if active_tab == 'semua' else ''

        qs_semua = RiwayatKlasifikasi.objects.select_related('user').all()
        if genre_filter_semua:
            qs_semua = qs_semua.filter(genre_prediksi__iexact=genre_filter_semua)
        if user_filter_semua:
            qs_semua = qs_semua.filter(user__username__icontains=user_filter_semua)
        if date_from_semua:
            try:
                qs_semua = qs_semua.filter(waktu_klasifikasi__date__gte=date_from_semua)
            except Exception:
                pass
        if date_to_semua:
            try:
                qs_semua = qs_semua.filter(waktu_klasifikasi__date__lte=date_to_semua)
            except Exception:
                pass

        for r in qs_semua:
            try:
                probabilities = json.loads(r.probabilities_json) if r.probabilities_json else []
            except Exception:
                probabilities = []
            genre_detail = GENRE_INFO.get(r.genre_prediksi.lower(), {})
            riwayat_semua.append({
                'id': r.id_riwayat,
                'username': r.user.username,
                'nama_file': r.nama_file,
                'ukuran_file': r.ukuran_file,
                'format_file': r.format_file,
                'genre_prediksi': r.genre_prediksi,
                'genre_icon': r.genre_icon,
                'genre_color': genre_detail.get('color', '#888'),
                'confidence': r.confidence,
                'waktu': r.waktu_klasifikasi,
                'probabilities_json': json.dumps(probabilities),
            })

        total_semua = RiwayatKlasifikasi.objects.count()
        for r in RiwayatKlasifikasi.objects.all():
            genre_stats_semua[r.genre_prediksi] = genre_stats_semua.get(r.genre_prediksi, 0) + 1
        raw_genres_semua = RiwayatKlasifikasi.objects.values_list('genre_prediksi', flat=True)
        all_genres_semua = sorted(list(set(g.strip().lower() for g in raw_genres_semua if g)))

    context = {
        'role': role,
        'is_advanced': is_advanced,
        'active_tab': active_tab,
        # Tab saya
        'riwayat_saya': riwayat_saya,
        'total_saya': total_saya,
        'genre_filter_saya': genre_filter_saya,
        'all_genres_saya': all_genres_saya,
        'genre_stats_saya': json.dumps(genre_stats_saya),
        'jumlah_ditampilkan_saya': len(riwayat_saya),
        # Tab semua
        'riwayat_semua': riwayat_semua,
        'total_semua': total_semua,
        'jumlah_ditampilkan_semua': len(riwayat_semua),
        'genre_filter_semua': genre_filter_semua,
        'user_filter_semua': user_filter_semua,
        'date_from_semua': date_from_semua,
        'date_to_semua': date_to_semua,
        'all_genres_semua': all_genres_semua,
        'genre_stats_semua': json.dumps(genre_stats_semua),
    }
    return render(request, 'Klasifikasi/riwayat.html', context)

@login_and_active_required
def hapus_riwayat(request, pk):
    """Hapus satu item riwayat milik sendiri atau semua riwayat jika admin/pengelola."""
    from .models import RiwayatKlasifikasi
    from .decorators import get_user_role
    
    if request.method != 'POST':
        return redirect('Klasifikasi:riwayat')
    try:
        role = get_user_role(request.user)
        is_advanced = role in ('admin', 'pengelola_studio')
        
        if is_advanced:
            item = RiwayatKlasifikasi.objects.get(pk=pk)
        else:
            item = RiwayatKlasifikasi.objects.get(pk=pk, user=request.user)
            
        nama = item.nama_file
        item.delete()
        messages.success(request, f'Riwayat "{nama}" berhasil dihapus.')
    except RiwayatKlasifikasi.DoesNotExist:
        messages.error(request, 'Item tidak ditemukan atau bukan milik Anda.')
    except Exception as e:
        messages.error(request, f'Gagal menghapus: {str(e)}')
    return redirect('Klasifikasi:riwayat')


@login_and_active_required
def hapus_semua_riwayat(request):
    """Hapus semua riwayat user sendiri."""
    from .models import RiwayatKlasifikasi
    if request.method == 'POST':
        RiwayatKlasifikasi.objects.filter(user=request.user).delete()
        messages.success(request, 'Semua riwayat berhasil dihapus.')
    return redirect('Klasifikasi:riwayat')

# ═══════════════════════════════════════════════════════════════════════
# LAPORAN — generate & download
# ═══════════════════════════════════════════════════════════════════════

@login_and_active_required
def laporan_view(request):
    """
    Manajemen laporan.
    - pengguna_studio: hanya laporan milik sendiri, tanpa tab.
    - admin / pengelola_studio: 2 tab — "Laporan Saya" dan "Semua Laporan".
      Tab aktif dikontrol via ?tab=semua (default: 'saya').
    """
    from .models import RiwayatKlasifikasi, Laporan

    role = get_user_role(request.user)
    is_advanced = role in ('admin', 'pengelola_studio')

    # ── Tab aktif ──
    active_tab = request.GET.get('tab', 'saya')
    if not is_advanced:
        active_tab = 'saya'  # paksa tab saya untuk pengguna biasa

    # ── Riwayat laporan yang sudah di-download (tb_laporan) ──
    if active_tab == 'semua' and is_advanced:
        laporan_list = Laporan.objects.select_related('id_user').all()
    else:
        laporan_list = Laporan.objects.filter(id_user=request.user)

    # ── Filter preview (tb_riwayat_klasifikasi) ──
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    genre_filter = request.GET.get('genre', '')
    user_filter = request.GET.get('user', '') if active_tab == 'semua' else ''
    sudah_filter = date_from or date_to or genre_filter or user_filter

    if active_tab == 'semua' and is_advanced:
        qs = RiwayatKlasifikasi.objects.select_related('user').all()
    else:
        qs = RiwayatKlasifikasi.objects.filter(user=request.user)

    if date_from:
        try:
            qs = qs.filter(waktu_klasifikasi__date__gte=date_from)
        except Exception:
            pass
    if date_to:
        try:
            qs = qs.filter(waktu_klasifikasi__date__lte=date_to)
        except Exception:
            pass
    if genre_filter:
        qs = qs.filter(genre_prediksi__iexact=genre_filter)
    if user_filter and active_tab == 'semua':
        qs = qs.filter(user__username__icontains=user_filter)

    # Buat preview list
    preview = []
    for idx, r in enumerate(qs, 1):
        genre_detail = GENRE_INFO.get(r.genre_prediksi.lower(), {})
        preview.append({
            'no': idx,
            'username': r.user.username,
            'nama_file': r.nama_file,
            'format_file': r.format_file,
            'genre_prediksi': r.genre_prediksi,
            'genre_icon': r.genre_icon,
            'genre_color': genre_detail.get('color', '#888'),
            'confidence': r.confidence,
            'waktu': r.waktu_klasifikasi,
        })

    # Genre list untuk dropdown filter
    if active_tab == 'semua' and is_advanced:
        raw_genres_laporan = RiwayatKlasifikasi.objects.values_list('genre_prediksi', flat=True)
    else:
        raw_genres_laporan = RiwayatKlasifikasi.objects.filter(user=request.user).values_list('genre_prediksi', flat=True)
    all_genres = sorted(list(set(g.strip().lower() for g in raw_genres_laporan if g)))

    context = {
        'role': role,
        'is_advanced': is_advanced,
        'active_tab': active_tab,
        'laporan_list': laporan_list,
        'preview': preview,
        'total_preview': len(preview),
        'sudah_filter': sudah_filter,
        'date_from': date_from,
        'date_to': date_to,
        'genre_filter': genre_filter,
        'user_filter': user_filter,
        'all_genres': all_genres,
    }
    return render(request, 'Klasifikasi/laporan.html', context)



def _build_riwayat_for_export(request, date_from_str, date_to_str, active_tab, role, genre_str='', user_str=''):
    """Helper: kembalikan queryset riwayat sesuai tab aktif, tanggal, genre, dan user."""
    from .models import RiwayatKlasifikasi
    if active_tab == 'semua' and role in ('admin', 'pengelola_studio'):
        qs = RiwayatKlasifikasi.objects.select_related('user').all()
        if user_str:
            qs = qs.filter(user__username__icontains=user_str)
    else:
        qs = RiwayatKlasifikasi.objects.filter(user=request.user)

    if date_from_str:
        try:
            qs = qs.filter(waktu_klasifikasi__date__gte=date_from_str)
        except Exception:
            pass
    if date_to_str:
        try:
            qs = qs.filter(waktu_klasifikasi__date__lte=date_to_str)
        except Exception:
            pass
    if genre_str:
        qs = qs.filter(genre_prediksi__iexact=genre_str)
    return qs


def _save_laporan_record(request, judul, date_from_str, date_to_str, fmt):
    """Helper: simpan record ke tb_laporan."""
    from .models import Laporan
    try:
        Laporan.objects.create(
            id_user=request.user,
            judul_laporan=judul,
            periode_awal=date_from_str if date_from_str else None,
            periode_akhir=date_to_str if date_to_str else None,
            format_ekspor=fmt,
        )
    except Exception as e:
        print(f"[WARNING] Gagal menyimpan record Laporan: {e}")


@login_and_active_required
def download_laporan_pdf(request):
    """Generate dan download laporan riwayat sebagai PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return HttpResponse(
            'Library ReportLab belum terinstall. Jalankan: pip install reportlab',
            status=500
        )

    role = get_user_role(request.user)
    active_tab = request.GET.get('tab', 'saya')
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    genre_str = request.GET.get('genre', '')
    user_str = request.GET.get('user', '') if active_tab == 'semua' else ''
    qs = _build_riwayat_for_export(request, date_from_str, date_to_str, active_tab, role, genre_str, user_str)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)

    elements = []
    elements.append(Paragraph('Studio Dungeon Limo', title_style))
    elements.append(Paragraph('Laporan Riwayat Klasifikasi Genre Musik', sub_style))

    periode = ''
    if date_from_str and date_to_str:
        periode = f'Periode: {date_from_str} s.d. {date_to_str}'
    elif date_from_str:
        periode = f'Sejak: {date_from_str}'
    elif date_to_str:
        periode = f'Sampai: {date_to_str}'
    if periode:
        elements.append(Paragraph(periode, sub_style))

    elements.append(Paragraph(f'Dicetak: {datetime.now().strftime("%d %B %Y %H:%M")}', sub_style))
    elements.append(Spacer(1, 0.5*cm))

    header = ['No', 'Pengguna', 'Nama File', 'Format', 'Genre Prediksi', 'Confidence', 'Waktu Klasifikasi']
    data = [header]
    for idx, r in enumerate(qs, 1):
        row_user = r.user.username

        data.append([
            str(idx),
            row_user,
            r.nama_file[:40],
            r.format_file,
            r.genre_prediksi,
            f"{r.confidence:.1f}%" if r.confidence else 'N/A',
            r.waktu_klasifikasi.strftime('%d/%m/%Y %H:%M') if r.waktu_klasifikasi else '-',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6d28d9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f0ff')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)

    pdf_data = buffer.getvalue()
    buffer.close()

    judul = f"Laporan Klasifikasi PDF {datetime.now().strftime('%Y%m%d_%H%M%S')}"
    _save_laporan_record(request, judul, date_from_str, date_to_str, 'pdf')

    response = HttpResponse(pdf_data, content_type='application/pdf')
    filename = f"laporan_klasifikasi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_and_active_required
def download_laporan_excel(request):
    """Generate dan download laporan riwayat sebagai Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'Library openpyxl belum terinstall. Jalankan: pip install openpyxl',
            status=500
        )

    role = get_user_role(request.user)
    active_tab = request.GET.get('tab', 'saya')
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    genre_str = request.GET.get('genre', '')
    user_str = request.GET.get('user', '') if active_tab == 'semua' else ''
    qs = _build_riwayat_for_export(request, date_from_str, date_to_str, active_tab, role, genre_str, user_str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Riwayat Klasifikasi'

    # Judul
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Studio Dungeon Limo - Laporan Riwayat Klasifikasi Genre Musik'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='6D28D9', end_color='6D28D9', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    if date_from_str or date_to_str:
        ws.merge_cells('A2:G2')
        periode = ''
        if date_from_str and date_to_str:
            periode = f'Periode: {date_from_str} s.d. {date_to_str}'
        elif date_from_str:
            periode = f'Sejak: {date_from_str}'
        else:
            periode = f'Sampai: {date_to_str}'
        ws['A2'] = periode
        ws['A2'].alignment = Alignment(horizontal='center')

    header_row = 3
    headers = ['No', 'Pengguna', 'Nama File', 'Format', 'Genre Prediksi', 'Confidence (%)', 'Waktu Klasifikasi']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, r in enumerate(qs, 1):
        row_num = header_row + idx
        row_user = r.user.username
        row_data = [
            idx,
            row_user,
            r.nama_file,
            r.format_file,
            r.genre_prediksi,
            round(r.confidence, 2) if r.confidence else 'N/A',
            r.waktu_klasifikasi.strftime('%d/%m/%Y %H:%M') if r.waktu_klasifikasi else '-',
        ]
        fill_color = 'F3F0FF' if idx % 2 == 0 else 'FFFFFF'
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center' if col in (1, 4, 6) else 'left', vertical='center')

    # Column widths
    col_widths = [6, 18, 40, 10, 18, 16, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_data = output.read()
    output.close()

    judul = f"Laporan Klasifikasi Excel {datetime.now().strftime('%Y%m%d_%H%M%S')}"
    _save_laporan_record(request, judul, date_from_str, date_to_str, 'xlsx')

    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan_klasifikasi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═══════════════════════════════════════════════════════════════════════
# CETAK LAPORAN — versi print
# ═══════════════════════════════════════════════════════════════════════

@login_and_active_required
def cetak_laporan(request):
    """Halaman cetak laporan riwayat klasifikasi (print view).
    Bisa difilter berdasarkan tanggal dan genre.
    Admin/pengelola bisa cetak semua data.
    """


    role = get_user_role(request.user)
    active_tab = request.GET.get('tab', 'saya')
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    genre_filter  = request.GET.get('genre', '')
    user_filter   = request.GET.get('user', '') if active_tab == 'semua' else ''

    qs = _build_riwayat_for_export(request, date_from_str, date_to_str, active_tab, role, genre_filter, user_filter)

    riwayat_parsed = []
    genre_stats = {}
    for idx, r in enumerate(qs, 1):
        genre_color = GENRE_INFO.get(r.genre_prediksi.lower(), {}).get('color', '#888')
        username = r.user.username
        riwayat_parsed.append({
            'no': idx,
            'username': username,
            'nama_file': r.nama_file,
            'ukuran_file': r.ukuran_file,
            'format_file': r.format_file,
            'genre_prediksi': r.genre_prediksi,
            'genre_icon': r.genre_icon,
            'genre_color': genre_color,
            'confidence': f"{r.confidence:.1f}%" if r.confidence else "N/A",
            'waktu': r.waktu_klasifikasi,
        })
        genre_stats[r.genre_prediksi] = genre_stats.get(r.genre_prediksi, 0) + 1

    # Buat label periode untuk header print
    if date_from_str and date_to_str:
        periode = f"{date_from_str} s.d. {date_to_str}"
    elif date_from_str:
        periode = f"Sejak {date_from_str}"
    elif date_to_str:
        periode = f"Sampai {date_to_str}"
    else:
        periode = "Semua Periode"

    context = {
        'riwayat': riwayat_parsed,
        'total': len(riwayat_parsed),
        'genre_filter': genre_filter,
        'genre_stats': sorted(genre_stats.items(), key=lambda x: x[1], reverse=True),
        'tanggal_cetak': datetime.now().strftime('%d %B %Y %H:%M'),
        'user': request.user,
        'role': role,
        'show_username_col': active_tab == 'semua' and role in ('admin', 'pengelola_studio'),
        'periode': periode,
        'date_from': date_from_str,
        'date_to': date_to_str,
    }
    return render(request, 'Klasifikasi/laporan_print.html', context)


# ═══════════════════════════════════════════════════════════════════════
# USER MANAGEMENT — admin only
# ═══════════════════════════════════════════════════════════════════════

@role_required('admin')
def user_management_view(request):
    """Manajemen pengguna — hanya admin."""

    users = CustomUser.objects.select_related('profile').all().order_by('date_joined')
    user_data = []
    admin_count = 0
    pengelola_count = 0
    pengguna_aktif_count = 0
    for u in users:
        try:
            profile = u.profile
            role = profile.role
            is_active_profile = profile.is_active
        except Exception:
            role = 'pengguna_studio'
            is_active_profile = True
            
        if u.is_superuser or role == 'admin':
            admin_count += 1
        elif role == 'pengelola_studio':
            pengelola_count += 1
        
        if is_active_profile:
            pengguna_aktif_count += 1
            
        user_data.append({
            'id_user': u.id_user,
            'username': u.username,
            'email': u.email,
            'first_name': u.first_name,
            'role': role,
            'is_active': is_active_profile,
            'date_joined': u.date_joined,
            'is_superuser': u.is_superuser,
        })

    context = {
        'users': user_data,
        'total_users': len(user_data),
        'admin_count': admin_count,
        'pengelola_count': pengelola_count,
        'pengguna_aktif_count': pengguna_aktif_count,
        'role_choices': [('admin', 'Admin'), ('pengelola_studio', 'Pengelola Studio'), ('pengguna_studio', 'Pengguna Studio')],
    }
    return render(request, 'Klasifikasi/user_management.html', context)


@role_required('admin')
def ubah_role_user(request, user_id):
    """Ubah role pengguna — hanya admin."""
    from .models import UserProfile
    if request.method != 'POST':
        return redirect('Klasifikasi:user_management')
    target_user = get_object_or_404(CustomUser, pk=user_id)
    if target_user == request.user:
        messages.error(request, 'Anda tidak dapat mengubah role Anda sendiri.')
        return redirect('Klasifikasi:user_management')
    new_role = request.POST.get('role', 'pengguna_studio')
    valid_roles = ['admin', 'pengelola_studio', 'pengguna_studio']
    if new_role not in valid_roles:
        messages.error(request, 'Role tidak valid.')
        return redirect('Klasifikasi:user_management')
    profile, _ = UserProfile.objects.get_or_create(user=target_user)
    profile.role = new_role
    profile.save()
    messages.success(request, f'Role pengguna "{target_user.username}" diubah menjadi {profile.get_role_display()}.')
    return redirect('Klasifikasi:user_management')


@role_required('admin')
def hapus_user(request, user_id):
    """Hapus akun pengguna secara permanen — hanya admin, dan hanya jika user nonaktif."""
    if request.method != 'POST':
        return redirect('Klasifikasi:user_management')
    target_user = get_object_or_404(CustomUser, pk=user_id)
    if target_user == request.user:
        messages.error(request, 'Anda tidak dapat menghapus akun Anda sendiri.')
        return redirect('Klasifikasi:user_management')
    
    # Cek status aktif dari profile
    try:
        is_active_profile = target_user.profile.is_active
    except Exception:
        is_active_profile = True
    
    if is_active_profile:
        messages.error(request, f'Akun "{target_user.username}" masih aktif dan tidak dapat dihapus. Nonaktifkan akun terlebih dahulu.')
        return redirect('Klasifikasi:user_management')
    
    username = target_user.username
    target_user.delete()
    messages.success(request, f'Akun "{username}" berhasil dihapus secara permanen.')
    return redirect('Klasifikasi:user_management')


@role_required('admin')
def edit_user(request, user_id):
    """Edit data pengguna dan reset password — hanya admin."""
    if request.method != 'POST':
        return redirect('Klasifikasi:user_management')
    target_user = get_object_or_404(CustomUser, pk=user_id)
    
    username = request.POST.get('username', '').strip()
    email = request.POST.get('email', '').strip()
    first_name = request.POST.get('first_name', '').strip()
    password = request.POST.get('password', '')
    
    if username:
        # Cek apakah username sudah dipakai orang lain
        if CustomUser.objects.filter(username=username).exclude(pk=user_id).exists():
            messages.error(request, f'Username "{username}" sudah digunakan oleh pengguna lain.')
            return redirect('Klasifikasi:user_management')
        target_user.username = username
        
    target_user.email = email
    target_user.first_name = first_name
    
    if password:
        target_user.set_password(password)
        
    target_user.save()
    messages.success(request, f'Data pengguna "{target_user.username}" berhasil diperbarui.')
    return redirect('Klasifikasi:user_management')


@role_required('admin')
def toggle_active_user(request, user_id):
    """Toggle status aktif/nonaktif pengguna — hanya admin."""
    from .models import UserProfile
    if request.method != 'POST':
        return redirect('Klasifikasi:user_management')
    target_user = get_object_or_404(CustomUser, pk=user_id)
    if target_user == request.user:
        messages.error(request, 'Anda tidak dapat mengubah status akun Anda sendiri.')
        return redirect('Klasifikasi:user_management')
    
    profile, _ = UserProfile.objects.get_or_create(user=target_user)
    profile.is_active = not profile.is_active
    profile.save()
    
    status_text = 'diaktifkan' if profile.is_active else 'dinonaktifkan'
    messages.success(request, f'Akun "{target_user.username}" berhasil {status_text}.')
    return redirect('Klasifikasi:user_management')
